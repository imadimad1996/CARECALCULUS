"""
CareCalculus Hospital Email Harvester
=====================================
Crawls USA, Canada, and UK hospital directories to extract
decision-maker contact emails (CMO, CIO, Medical Director, etc.)
for CareCalculus B2B marketing outreach.

Usage:
    pip install crawl4ai openpyxl
    python crawler.py

Output:
    output/hospital_leads.csv   — All discovered leads
    output/hospital_leads.xlsx  — Formatted Excel workbook
    output/run_log.txt          — Detailed crawl log
"""

import asyncio
import csv
import re
import random
import logging
from datetime import datetime
from pathlib import Path

from crawl4ai import AsyncWebCrawler, BrowserConfig, CrawlerRunConfig

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
OUTPUT_DIR = Path(__file__).parent.parent.parent / "hospital_emails"
OUTPUT_DIR.mkdir(exist_ok=True)

LOG_FILE   = OUTPUT_DIR / "run_log.txt"
CSV_FILE   = OUTPUT_DIR / "hospital_leads.csv"
XLSX_FILE  = OUTPUT_DIR / "hospital_leads.xlsx"

MAX_CONCURRENT = 3       # Parallel crawls (polite)
DELAY_MIN      = 1.0     # Seconds between batches (min)
DELAY_MAX      = 2.5     # Seconds between batches (max)
PAGE_TIMEOUT   = 20_000  # 20s per page (was 45s)
HOSPITAL_TIMEOUT = 60   # Max seconds total per hospital

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# TARGET HOSPITALS  (direct contact page URLs — highest value)
# ─────────────────────────────────────────────────────────────────────────────
TARGET_HOSPITALS = [
    # ── USA ──────────────────────────────────────────────────────────────────
    {"country": "USA", "name": "Mayo Clinic",                     "url": "https://www.mayoclinic.org/about-mayo-clinic/contact"},
    {"country": "USA", "name": "Cleveland Clinic",                "url": "https://my.clevelandclinic.org/departments/marketing-communications"},
    {"country": "USA", "name": "Johns Hopkins Hospital",          "url": "https://www.hopkinsmedicine.org/contact-us"},
    {"country": "USA", "name": "Massachusetts General Hospital",  "url": "https://www.massgeneral.org/contact"},
    {"country": "USA", "name": "UCSF Medical Center",            "url": "https://www.ucsfhealth.org/contact-us"},
    {"country": "USA", "name": "NewYork-Presbyterian",           "url": "https://www.nyp.org/contact-us"},
    {"country": "USA", "name": "Cedars-Sinai",                   "url": "https://www.cedars-sinai.org/contact-us.html"},
    {"country": "USA", "name": "UPMC",                           "url": "https://www.upmc.com/contact-us"},
    {"country": "USA", "name": "Duke University Hospital",        "url": "https://www.dukehealth.org/contact"},
    {"country": "USA", "name": "Northwestern Memorial Hospital",  "url": "https://www.nm.org/contact-us"},
    {"country": "USA", "name": "Brigham and Women's Hospital",   "url": "https://www.brighamandwomens.org/contact-us"},
    {"country": "USA", "name": "Stanford Health Care",           "url": "https://stanfordhealthcare.org/contact-us"},
    {"country": "USA", "name": "Houston Methodist Hospital",      "url": "https://www.houstonmethodist.org/contact-us"},
    {"country": "USA", "name": "Penn Medicine",                  "url": "https://www.pennmedicine.org/contact"},
    {"country": "USA", "name": "Vanderbilt University Medical",  "url": "https://www.vumc.org/main/contact"},
    {"country": "USA", "name": "OHSU Hospital",                  "url": "https://www.ohsu.edu/contact"},
    {"country": "USA", "name": "Rush University Medical",        "url": "https://www.rush.edu/contact-rush"},
    {"country": "USA", "name": "Emory Healthcare",               "url": "https://www.emoryhealthcare.org/contact-emory"},
    {"country": "USA", "name": "Yale New Haven Hospital",        "url": "https://www.ynhh.org/contact"},
    {"country": "USA", "name": "UC San Diego Health",            "url": "https://health.ucsd.edu/contact/pages/default.aspx"},
    {"country": "USA", "name": "Indiana University Health",      "url": "https://iuhealth.org/contact-us"},
    {"country": "USA", "name": "UCHealth Colorado",              "url": "https://www.uchealth.org/contact"},
    {"country": "USA", "name": "Ochsner Health Louisiana",       "url": "https://www.ochsner.org/contact-us"},
    {"country": "USA", "name": "Wake Forest Baptist Health",     "url": "https://www.wakehealth.edu/contact-us"},
    {"country": "USA", "name": "Baylor Scott & White Health",    "url": "https://www.bswhealth.com/contact-us"},
    {"country": "USA", "name": "Mount Sinai Health System",      "url": "https://www.mountsinai.org/about/contact"},
    {"country": "USA", "name": "NYU Langone Health",             "url": "https://nyulangone.org/contact-us"},
    {"country": "USA", "name": "UCLA Health",                    "url": "https://www.uclahealth.org/contact-us"},
    {"country": "USA", "name": "Memorial Sloan Kettering",       "url": "https://www.mskcc.org/contact"},
    {"country": "USA", "name": "Children's Hosp of Philadelphia", "url": "https://www.chop.edu/contact-us"},
    {"country": "USA", "name": "Boston Children's Hospital",     "url": "https://www.childrenshospital.org/contact-us"},
    {"country": "USA", "name": "Dana-Farber Cancer Institute",   "url": "https://www.dana-farber.org/contact-us"},
    {"country": "USA", "name": "MD Anderson Cancer Center",      "url": "https://www.mdanderson.org/about-md-anderson/contact-us.html"},
    {"country": "USA", "name": "Ascension Health",               "url": "https://healthcare.ascension.org/contact-us"},
    {"country": "USA", "name": "CommonSpirit Health",            "url": "https://www.commonspirit.org/contact-us"},
    {"country": "USA", "name": "Providence Health",              "url": "https://www.providence.org/contact-us"},
    {"country": "USA", "name": "Trinity Health",                 "url": "https://www.trinity-health.org/contact-us"},
    {"country": "USA", "name": "Advocate Health",                "url": "https://www.advocatehealth.com/contact-us"},
    {"country": "USA", "name": "Intermountain Health",           "url": "https://intermountainhealthcare.org/about/contact-us"},
    {"country": "USA", "name": "Banner Health",                  "url": "https://www.bannerhealth.com/contact-us"},
    {"country": "USA", "name": "Kaiser Permanente",              "url": "https://healthy.kaiserpermanente.org/support"},
    {"country": "USA", "name": "Geisinger Health System",        "url": "https://www.geisinger.org/about-geisinger/contact-us"},
    {"country": "USA", "name": "Scripps Health",                 "url": "https://www.scripps.org/contact_us"},
    {"country": "USA", "name": "Jefferson Health",               "url": "https://www.jeffersonhealth.org/contact-us"},
    {"country": "USA", "name": "Northwell Health",               "url": "https://www.northwell.org/about/contact-us"},
    {"country": "USA", "name": "Hackensack Meridian Health",    "url": "https://www.hackensackmeridianhealth.org/en/contact-us"},
    {"country": "USA", "name": "RWJBarnabas Health",            "url": "https://www.rwjbh.org/contact-us/"},
    {"country": "USA", "name": "Beth Israel Deaconess",          "url": "https://www.bidmc.org/contact-us"},
    {"country": "USA", "name": "Stony Brook Medicine",           "url": "https://www.stonybrookmedicine.edu/contact-us"},
    {"country": "USA", "name": "UChicago Medicine",              "url": "https://www.uchicagomedicine.org/contact-us"},
    {"country": "USA", "name": "Corewell Health",                "url": "https://www.corewellhealth.org/contact-us"},
    {"country": "USA", "name": "Henry Ford Health",              "url": "https://www.henryford.com/contact-us"},
    {"country": "USA", "name": "OhioState Wexner Medical Center", "url": "https://wexnermedical.osu.edu/utility/contact-us"},
    {"country": "USA", "name": "UR Medicine Rochester",          "url": "https://www.urmc.rochester.edu/contact.aspx"},
    {"country": "USA", "name": "Inova Health System",            "url": "https://www.inova.org/contact-us"},
    {"country": "USA", "name": "VCU Health",                     "url": "https://www.vcuhealth.org/contact-us"},
    {"country": "USA", "name": "UVA Health",                     "url": "https://uvahealth.com/services/contact-us"},
    {"country": "USA", "name": "MUSC Health",                    "url": "https://muschealth.org/about-us/contact-us"},
    {"country": "USA", "name": "UNC Health",                     "url": "https://www.unchealthcare.org/contact-us"},
    {"country": "USA", "name": "UF Health Shands",               "url": "https://ufhealth.org/contact-us"},
    {"country": "USA", "name": "Moffitt Cancer Center",          "url": "https://www.moffitt.org/about-moffitt/contact-us"},
    {"country": "USA", "name": "Tampa General Hospital",         "url": "https://www.tgh.org/contact-us"},
    {"country": "USA", "name": "Baptist Health South Florida",   "url": "https://baptisthealth.net/contact-us"},
    {"country": "USA", "name": "UAB Medicine",                   "url": "https://www.uabmedicine.org/contact-us"},
    {"country": "USA", "name": "St. Jude Children's",            "url": "https://www.stjude.org/contact-us.html"},
    {"country": "USA", "name": "Barnes-Jewish BJC Health",       "url": "https://www.barnesjewish.org/Contact-Us"},
    {"country": "USA", "name": "University of Kansas Health",    "url": "https://www.kansashealthsystem.com/contact-us"},
    {"country": "USA", "name": "Nebraska Medicine",              "url": "https://www.nebraskamed.com/contact-us"},
    {"country": "USA", "name": "Sanford Health",                 "url": "https://www.sanfordhealth.org/contact-us"},
    {"country": "USA", "name": "M Health Fairview",              "url": "https://mhealthfairview.org/contact-us"},
    {"country": "USA", "name": "UW Health Wisconsin",            "url": "https://www.uwhealth.org/contact-us"},
    {"country": "USA", "name": "Froedtert & MCW",                "url": "https://www.froedtert.com/contact"},
    {"country": "USA", "name": "City of Hope",                   "url": "https://www.cityofhope.org/contact-us"},
    {"country": "USA", "name": "Sutter Health",                  "url": "https://www.sutterhealth.org/contact"},
    {"country": "USA", "name": "UC Davis Health",                "url": "https://health.ucdavis.edu/contactus/"},
    {"country": "USA", "name": "UC Irvine Health",               "url": "https://www.ucihealth.org/contact-us"},

    # ── CANADA ────────────────────────────────────────────────────────────────
    {"country": "Canada", "name": "University Health Network",   "url": "https://www.uhn.ca/corporate/contact"},
    {"country": "Canada", "name": "McGill University Health Centre","url": "https://www.muhc.ca/contact"},
    {"country": "Canada", "name": "Vancouver Coastal Health",    "url": "https://www.vch.ca/contact"},
    {"country": "Canada", "name": "Ottawa Hospital",             "url": "https://www.ottawahospital.on.ca/en/contact-us"},
    {"country": "Canada", "name": "Sunnybrook Health Sciences",  "url": "https://sunnybrook.ca/content/?page=contact-us"},
    {"country": "Canada", "name": "Alberta Health Services",     "url": "https://www.albertahealthservices.ca/contact.aspx"},
    {"country": "Canada", "name": "QEII Health Sciences Centre", "url": "https://www.nshealth.ca/contact"},
    {"country": "Canada", "name": "Kingston Health Sciences",    "url": "https://kingstonhsc.ca/contact"},
    {"country": "Canada", "name": "Hamilton Health Sciences",    "url": "https://www.hamiltonhealthsciences.ca/contact"},
    {"country": "Canada", "name": "London Health Sciences",      "url": "https://www.lhsc.on.ca/contact-us"},
    {"country": "Canada", "name": "Winnipeg Health Sciences",    "url": "https://www.wrha.mb.ca/contact"},
    {"country": "Canada", "name": "CIUSSS de l'Est-de-l'Île-de-Montréal","url": "https://ciusss-estmtl.gouv.qc.ca/contact"},
    {"country": "Canada", "name": "Sinai Health Toronto",        "url": "https://www.sinaihealth.ca/contact-us/"},
    {"country": "Canada", "name": "Unity Health Toronto",        "url": "https://unityhealth.to/contact-us/"},
    {"country": "Canada", "name": "SickKids Toronto",            "url": "https://www.sickkids.ca/en/contact-us/"},
    {"country": "Canada", "name": "CAMH Toronto",                "url": "https://www.camh.ca/en/your-care/contact-us"},
    {"country": "Canada", "name": "Trillium Health Partners",    "url": "https://www.trilliumhealthpartners.ca/contactus/Pages/default.aspx"},
    {"country": "Canada", "name": "Fraser Health BC",            "url": "https://www.fraserhealth.ca/contact-us"},
    {"country": "Canada", "name": "Interior Health BC",          "url": "https://www.interiorhealth.ca/about-us/contact-us"},
    {"country": "Canada", "name": "Covenant Health Alberta",     "url": "https://www.covenanthealth.ca/contact-us"},
    {"country": "Canada", "name": "CHUM Montreal",               "url": "https://www.chumontreal.qc.ca/nous-joindre"},
    {"country": "Canada", "name": "CHU Sainte-Justine",          "url": "https://www.chusj.org/fr/Nous-joindre"},

    # ── UK ───────────────────────────────────────────────────────────────────
    {"country": "UK", "name": "Royal London Hospital",            "url": "https://www.bartshealth.nhs.uk/contact"},
    {"country": "UK", "name": "Guy's and St Thomas' NHS",         "url": "https://www.guysandstthomas.nhs.uk/contact-us"},
    {"country": "UK", "name": "King's College Hospital",          "url": "https://www.kch.nhs.uk/contact"},
    {"country": "UK", "name": "Imperial College Healthcare NHS",  "url": "https://www.imperial.nhs.uk/contact-us"},
    {"country": "UK", "name": "Manchester NHS Foundation Trust",  "url": "https://mft.nhs.uk/contact-us"},
    {"country": "UK", "name": "Leeds Teaching Hospitals NHS",     "url": "https://www.leedsth.nhs.uk/contact-us"},
    {"country": "UK", "name": "Oxford University Hospitals NHS",  "url": "https://www.ouh.nhs.uk/contact"},
    {"country": "UK", "name": "Cambridge University Hospitals",   "url": "https://www.cuh.nhs.uk/contact-us"},
    {"country": "UK", "name": "Newcastle Hospitals NHS",          "url": "https://www.newcastle-hospitals.nhs.uk/contact"},
    {"country": "UK", "name": "University College London NHS",    "url": "https://www.uclh.nhs.uk/contact-us"},
    {"country": "UK", "name": "Sheffield Teaching Hospitals",     "url": "https://www.sth.nhs.uk/patients/contact-us"},
    {"country": "UK", "name": "Bristol University NHS Trust",     "url": "https://www.uhbristol.nhs.uk/contact-us"},
    {"country": "UK", "name": "Royal Free London NHS",            "url": "https://www.royalfree.nhs.uk/contact-us"},
    {"country": "UK", "name": "Nottingham University Hospitals",  "url": "https://www.nuh.nhs.uk/contact"},
    {"country": "UK", "name": "Liverpool University Hospitals",   "url": "https://www.liverpoolft.nhs.uk/contact-us"},
    {"country": "UK", "name": "University Hospitals Birmingham",  "url": "https://www.uhb.nhs.uk/contact-us"},
    {"country": "UK", "name": "Southampton University NHS",       "url": "https://www.uhs.nhs.uk/contact"},
    {"country": "UK", "name": "Royal Marsden NHS Foundation",     "url": "https://www.royalmarsden.nhs.uk/contact-us"},
    {"country": "UK", "name": "Great Ormond Street NHS",          "url": "https://www.gosh.nhs.uk/contact-us/"},
    {"country": "UK", "name": "St George's University NHS",       "url": "https://www.stgeorges.nhs.uk/contact-us/"},
    {"country": "UK", "name": "Chelsea and Westminster NHS",      "url": "https://www.chelwest.nhs.uk/contact-us"},
    {"country": "UK", "name": "University Hospitals Sussex",      "url": "https://www.uhsussex.nhs.uk/contact-us/"},
    {"country": "UK", "name": "Northumbria Healthcare NHS",       "url": "https://www.northumbria.nhs.uk/contact-us"},
    {"country": "UK", "name": "University Hospitals of Leicester","url": "https://www.leicestershospitals.nhs.uk/contact-us/"},
    {"country": "UK", "name": "Derby and Burton NHS Trust",       "url": "https://www.uhdb.nhs.uk/contact-us"},
    {"country": "UK", "name": "Norfolk and Norwich NHS",          "url": "https://www.nnuh.nhs.uk/contact-us/"},
    {"country": "UK", "name": "East Suffolk & North Essex NHS",   "url": "https://www.esneft.nhs.uk/contact-us/"},
    {"country": "UK", "name": "NHS Greater Glasgow and Clyde",    "url": "https://www.nhsggc.scot/contact-us/"},
    {"country": "UK", "name": "NHS Lothian Edinburgh",            "url": "https://www.nhslothian.scot/contact-us/"},
    {"country": "UK", "name": "Cardiff and Vale Health Board",    "url": "https://cavuhb.nhs.wales/contact-us/"},
    {"country": "UK", "name": "Belfast Health & Social Care",     "url": "https://belfasttrust.hscni.net/contact-us/"},

    # ── AUSTRALIA & NEW ZEALAND ──────────────────────────────────────────────
    {"country": "Australia", "name": "Royal Melbourne Hospital",   "url": "https://www.thermh.org.au/contact-us"},
    {"country": "Australia", "name": "Alfred Health Melbourne",    "url": "https://www.alfredhealth.org.au/contact-us"},
    {"country": "Australia", "name": "St Vincent's Melbourne",     "url": "https://www.svhm.org.au/contact-us"},
    {"country": "Australia", "name": "Monash Health Melbourne",    "url": "https://monashhealth.org/contact-us/"},
    {"country": "Australia", "name": "Austin Health Melbourne",    "url": "https://www.austin.org.au/contact-us/"},
    {"country": "Australia", "name": "Royal Prince Alfred Sydney", "url": "https://www.slhd.nsw.gov.au/rpa/contact.html"},
    {"country": "Australia", "name": "St Vincent's Sydney",        "url": "https://www.svhs.org.au/contact-us"},
    {"country": "Australia", "name": "Royal North Shore Sydney",   "url": "https://www.nslhd.health.nsw.gov.au/Hospitals/RNSH/Pages/default.aspx"},
    {"country": "Australia", "name": "Royal Brisbane & Women's",   "url": "https://metronorth.health.qld.gov.au/rbwh/contact-us"},
    {"country": "Australia", "name": "Princess Alexandra Brisbane","url": "https://metro-south.health.qld.gov.au/princess-alexandra-hospital/contact-us"},
    {"country": "Australia", "name": "Royal Adelaide Hospital",    "url": "https://www.rah.sa.gov.au/contact-us"},
    {"country": "Australia", "name": "Sir Charles Gairdner Perth", "url": "https://www.nmahs.health.wa.gov.au/contact-us"},

    # ── IRELAND ──────────────────────────────────────────────────────────────
    {"country": "Ireland", "name": "St. James's Hospital Dublin",  "url": "https://www.stjames.ie/contact/"},
    {"country": "Ireland", "name": "Beaumont Hospital Dublin",     "url": "https://www.beaumont.ie/contact-us"},
    {"country": "Ireland", "name": "Mater Hospital Dublin",        "url": "https://www.mater.ie/contact/"},
    {"country": "Ireland", "name": "St. Vincent's Dublin",         "url": "https://www.stvincents.ie/contact-us/"},
    {"country": "Ireland", "name": "Galway University Hospitals",  "url": "https://www.saolta.ie/contact"},
]

# ─────────────────────────────────────────────────────────────────────────────
# EMAIL EXTRACTION & SCORING
# ─────────────────────────────────────────────────────────────────────────────
EMAIL_RE = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}",
    re.IGNORECASE,
)

# Match obfuscated emails: name [at] domain [dot] tld
OBFUSCATED_RE = re.compile(
    r"([a-zA-Z0-9._%+\-]+)\s*(?:\[at\]|\(at\)|\s+at\s+)\s*([a-zA-Z0-9.\-]+)\s*(?:\[dot\]|\(dot\)|\s+dot\s+)\s*([a-zA-Z]{2,})",
    re.IGNORECASE,
)

# mailto: href pattern in raw HTML
MAILTO_RE = re.compile(r'href=["\']mailto:([^"\'\s>?]+)', re.IGNORECASE)

HIGH_VALUE_EMAIL_KEYWORDS = [
    "cmo", "cio", "cco", "ceo", "director", "chief",
    "medical", "clinical", "digital", "info", "contact",
    "admin", "partnerships", "strategy", "innovation", "health",
    "press", "media", "communications", "procurement", "bd",
]

DECISION_MAKER_TITLES = [
    "chief medical officer", "chief information officer", "chief clinical officer",
    "medical director", "director of clinical", "vp clinical", "vice president",
    "chief executive", "cmo", "cio", "cco", "ceo", "head of", "partnerships",
    "innovation lead", "digital health", "it director", "director of informatics",
    "nursing director", "chief nursing officer", "chief of staff",
    "press office", "media contact", "communications director", "procurement",
]

SPAM_FILTER = [
    "noreply", "no-reply", "donotreply", "bounce", "mailer-daemon",
    "newsletter", "unsubscribe", "example.com", ".png", ".jpg", ".gif",
    "test@", "support@sentry", "wired.com", "schema.org", "w3.org",
    "sentry.io", "@2x", "jquery", "bootstrap", "example@", "@xyz.com",
    "@email.com", "abc@xyz.com",
]


def score_email(email: str, context: str = "") -> int:
    e   = email.lower()
    ctx = context.lower()
    if any(s in e for s in SPAM_FILTER):
        return 0
    # Reject clearly non-email strings
    if len(e) > 80 or e.count(".") > 5:
        return 0
    score = 10
    for kw in HIGH_VALUE_EMAIL_KEYWORDS:
        if kw in e:
            score += 5
    for title in DECISION_MAKER_TITLES:
        if title in ctx:
            score += 10
    return score


def extract_context(text: str, email: str, window: int = 250) -> str:
    idx = text.lower().find(email.lower())
    if idx < 0:
        return ""
    start = max(0, idx - window)
    end   = min(len(text), idx + len(email) + window)
    return " ".join(text[start:end].split())


def parse_emails(result, hospital_name: str, country: str) -> list[dict]:
    """Extract emails from multiple sources: raw HTML, mailto links, markdown, obfuscated."""
    seen, leads = set(), []

    raw_html = result.html or ""
    markdown  = result.markdown or ""

    # Source 1: mailto: href links in raw HTML (most reliable)
    mailto_emails = MAILTO_RE.findall(raw_html)

    # Source 2: plain email regex on raw HTML
    html_emails = EMAIL_RE.findall(raw_html)

    # Source 3: obfuscated emails (john [at] hospital [dot] nhs.uk)
    obfuscated = [
        f"{m[0]}@{m[1]}.{m[2]}"
        for m in OBFUSCATED_RE.findall(raw_html + " " + markdown)
    ]

    # Source 4: markdown plain text (fallback)
    md_emails = EMAIL_RE.findall(markdown)

    all_candidates = mailto_emails + html_emails + obfuscated + md_emails

    for email in all_candidates:
        email = email.lower().strip(".,;:()\"'")
        email = re.sub(r'^(?:u003e|\\u003e|//|<|>|mailto:)+', '', email, flags=re.IGNORECASE)
        email = email.strip(".,;:()\"'<> \t\n\r")
        # Strip query strings that sometimes get appended
        email = email.split("?")[0].split("&")[0]
        if email in seen or not email or "@" not in email or email.startswith("//") or "u003e" in email:
            continue
        seen.add(email)
        # Get context from raw HTML first (more text around it)
        context = extract_context(raw_html, email) or extract_context(markdown, email)
        s = score_email(email, context)
        if s == 0:
            continue
        leads.append({
            "score":      s,
            "hospital":   hospital_name,
            "country":    country,
            "email":      email,
            "context":    context[:220],
            "page_url":   result.url,
            "crawled_at": datetime.utcnow().isoformat(),
        })

    leads.sort(key=lambda x: -x["score"])
    return leads


# ─────────────────────────────────────────────────────────────────────────────
# AUTO CONTACT-PAGE DISCOVERY
# ─────────────────────────────────────────────────────────────────────────────
CONTACT_PATTERN = re.compile(
    r"(contact|about|leadership|team|staff|directory|administration"
    r"|board|governance|executives|management|people|find-us|reach-us"
    r"|get-in-touch|who-we-are|our-team)",
    re.IGNORECASE,
)


def find_contact_links(result) -> list[str]:
    from urllib.parse import urlparse
    parsed = urlparse(result.url)
    base   = f"{parsed.scheme}://{parsed.netloc}"
    links  = []
    for link in result.links.get("internal", []):
        href = link.get("href", "") if isinstance(link, dict) else str(link)
        if CONTACT_PATTERN.search(href) and href != result.url:
            full = href if href.startswith("http") else f"{base}{href}"
            links.append(full)
    return list(set(links))[:6]  # Max 6 sub-pages per hospital


# ─────────────────────────────────────────────────────────────────────────────
# CRAWL A SINGLE HOSPITAL
# ─────────────────────────────────────────────────────────────────────────────
async def crawl_hospital(crawler, hospital: dict) -> list[dict]:
    name, country, url = hospital["name"], hospital["country"], hospital["url"]
    all_leads = []
    config = CrawlerRunConfig(
        page_timeout=PAGE_TIMEOUT,
        remove_overlay_elements=True,
        excluded_tags=["script", "style", "nav"],
        exclude_external_links=True,
    )

    log.info(f"🏥 {name} ({country})")
    try:
        # Wrap entire hospital crawl in a hard timeout
        result = await asyncio.wait_for(
            crawler.arun(url=url, config=config),
            timeout=HOSPITAL_TIMEOUT
        )
        if not result.success:
            log.warning(f"  ❌ {result.error_message}")
            return all_leads

        leads = parse_emails(result, name, country)
        all_leads.extend(leads)
        log.info(f"  📧 {len(leads)} emails on main page")

        # Crawl discovered sub-pages
        sub_links = find_contact_links(result)
        if sub_links:
            try:
                sub_results = await asyncio.wait_for(
                    crawler.arun_many(
                        urls=sub_links, config=config, max_concurrent=2
                    ),
                    timeout=HOSPITAL_TIMEOUT
                )
                for sub in sub_results:
                    if sub and sub.success:
                        sub_leads = parse_emails(sub, name, country)
                        all_leads.extend(sub_leads)
                        if sub_leads:
                            log.info(f"    ↳ {sub.url.split('/')[-1]} → {len(sub_leads)} emails")
            except asyncio.TimeoutError:
                log.warning(f"  ⏱ Sub-pages timed out for {name}, skipping")

    except asyncio.TimeoutError:
        log.warning(f"  ⏱ Hard timeout for {name} — skipping")
    except Exception as exc:
        log.error(f"  💥 {exc}")

    return all_leads


# ─────────────────────────────────────────────────────────────────────────────
# SAVE OUTPUT
# ─────────────────────────────────────────────────────────────────────────────
def save_csv(leads: list[dict]):
    fields = ["score", "hospital", "country", "email", "context", "page_url", "crawled_at"]
    with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(leads)
    log.info(f"✅ CSV saved → {CSV_FILE}")


def save_xlsx(leads: list[dict]):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hospital Leads"

        headers = ["Score", "Hospital", "Country", "Email", "Context", "Source URL", "Crawled At"]
        hdr_fill = PatternFill("solid", fgColor="1A73E8")
        hdr_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        country_color = {"USA": "E8F4FD", "Canada": "E8F5E9", "UK": "FFF3E0"}
        for row, lead in enumerate(leads, 2):
            fill = PatternFill("solid", fgColor=country_color.get(lead["country"], "FFFFFF"))
            for col, key in enumerate(["score","hospital","country","email","context","page_url","crawled_at"], 1):
                cell = ws.cell(row=row, column=col, value=lead.get(key, ""))
                cell.fill = fill

        for col in range(1, len(headers) + 1):
            max_len = max(len(str(ws.cell(r, col).value or "")) for r in range(1, min(ws.max_row + 1, 60)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 65)

        wb.save(XLSX_FILE)
        log.info(f"✅ Excel saved → {XLSX_FILE}")
    except ImportError:
        log.warning("openpyxl not installed — skipping Excel. Run: pip install openpyxl")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
async def main():
    log.info("=" * 65)
    log.info("CareCalculus Hospital Email Harvester")
    log.info(f"Hospitals: {len(TARGET_HOSPITALS)} | Max concurrent: {MAX_CONCURRENT}")
    log.info("=" * 65)

    browser_config = BrowserConfig(
        headless=True,
        viewport_width=1440,
        viewport_height=900,
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/126.0.0.0 Safari/537.36"
        ),
    )

    all_leads: list[dict] = []
    seen_emails: set[str] = set()

    async with AsyncWebCrawler(config=browser_config) as crawler:
        for i in range(0, len(TARGET_HOSPITALS), MAX_CONCURRENT):
            batch = TARGET_HOSPITALS[i : i + MAX_CONCURRENT]
            batch_num = i // MAX_CONCURRENT + 1
            total_batches = -(-len(TARGET_HOSPITALS) // MAX_CONCURRENT)
            log.info(f"\n📦 Batch {batch_num}/{total_batches}")

            tasks = [crawl_hospital(crawler, h) for h in batch]
            results = await asyncio.gather(*tasks, return_exceptions=True)

            for r in results:
                if isinstance(r, list):
                    for lead in r:
                        if lead["email"] not in seen_emails:
                            seen_emails.add(lead["email"])
                            all_leads.append(lead)

            if all_leads:
                all_leads.sort(key=lambda x: (-x["score"], x["country"], x["hospital"]))
                save_csv(all_leads)
                save_xlsx(all_leads)

            delay = random.uniform(DELAY_MIN, DELAY_MAX)
            log.info(f"⏳ Sleeping {delay:.1f}s…")
            await asyncio.sleep(delay)

    # Sort: best score first, then by country and hospital name
    all_leads.sort(key=lambda x: (-x["score"], x["country"], x["hospital"]))

    if all_leads:
        save_csv(all_leads)
        save_xlsx(all_leads)
    else:
        log.warning("⚠️ No leads found — check connectivity or hospital URLs.")

    # ── Summary ───────────────────────────────────────────────────────────────
    log.info("\n" + "=" * 65)
    log.info(f"🎯 DONE — {len(all_leads)} unique leads harvested")
    by_country: dict[str, int] = {}
    for lead in all_leads:
        by_country[lead["country"]] = by_country.get(lead["country"], 0) + 1
    for country, count in sorted(by_country.items()):
        log.info(f"   {country}: {count} emails")
    high_value = [l for l in all_leads if l["score"] >= 20]
    log.info(f"\n⭐ High-value leads (score ≥ 20): {len(high_value)}")
    for lead in high_value[:15]:
        log.info(f"   [{lead['score']:>3}] {lead['email']:45s} {lead['hospital']} ({lead['country']})")
    log.info("=" * 65)
    log.info(f"\n📂 Output folder: {OUTPUT_DIR.resolve()}")


if __name__ == "__main__":
    asyncio.run(main())
